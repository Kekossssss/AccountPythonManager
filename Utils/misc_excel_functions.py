import openpyxl as opxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
from Utils.global_var import LANGUAGE_DICT

DICT_NUM_TO_ALPH = {0:'Z', 1:'A', 2:'B', 3:'C', 4:'D', 5:'E', 6:'F', 7:'G', 8:'H', 9:'I', 10:'J', 11:'K', 12:'L', 13:'M', 14:'N', 15:'O', 16:'P', 17:'Q', 18:'R', 19:'S', 20:'T', 21:'U', 22:'V', 23:'W', 24:'X', 25:'Y'}

## EXCEL DICT
def get_column(i):
    coef = int(i/26)
    if (coef == 0):
        a_1 = ""
    else:
        a_1 = DICT_NUM_TO_ALPH[coef]
    a_2 = DICT_NUM_TO_ALPH[i%26]
    a = a_1 + a_2
    return a

def compute_cell(i, j):
    ##TODO: WORK ONLY FOR THE 26*26 FIRST COLUMNS
    ##TODO: NEED FIXING FOR DICT VALUES AS Z COLUMNS HAVE AN OFFSET OF 26 (COLUMN 0 = Z, COLUMN 26 = AZ, ...)
    coef = int(i/26)
    if (coef == 0):
        a_1 = ""
    else:
        a_1 = DICT_NUM_TO_ALPH[coef]
    a_2 = DICT_NUM_TO_ALPH[i%26]
    a = a_1 + a_2
    b = str(j)
    return a+b


#############################################################
## GENERAL FUNCTIONS                                       ##
#############################################################
def apply_worksheet_background(worksheet, max_row = 100, max_col=100):
    worksheet.page_setup.fitToHeight = 1
    worksheet.page_setup.fitToWidth = 1
    ## Apply white background
    pf = PatternFill("solid", fgColor="FFFFFF")
    for i in range(1, max_row):
        for j in range(0, max_col):
            cell = worksheet[compute_cell(j,i)]
            cell.fill = pf

def set_columns_size(worksheet, nb_col=20):
    for j in range(2, nb_col):
        worksheet.column_dimensions[get_column(j)].width = 24

def apply_case_style(worksheet, row, col):
    background = PatternFill("solid", fgColor="A0A0A0")
    font = Font(color="000000", bold=True, size=20)
    border = Border(
        left   = Side(border_style='thick', color="000000"),
        right  = Side(border_style='thick', color="000000"),
        top    = Side(border_style='thick', color="000000"),
        bottom = Side(border_style='thick', color="000000")
    )
    cell = worksheet[compute_cell(col,row)]
    cell.fill = background
    cell.font = font
    cell.border = border
    cell.alignment = Alignment(wrap_text=True)

def apply_simple_vertical_table(worksheet, width=2, height=3, start_row=1, start_col=1, is_last_total=False, is_last_percent=False):
    data_background = [PatternFill("solid", fgColor="C0C0C0"), PatternFill("solid", fgColor="E0E0E0")]
    data_font = [Font(color="000000", bold=False, size=14), Font(color="000000", bold=True, size=14), Font(color="AA0000", bold=True, size=14), Font(color="00AA00", bold=True, size=14)]
    data_border = Border(
        left   = Side(border_style='thick', color="000000"),
        right  = Side(border_style='thick', color="000000"),
        top    = Side(border_style='thick', color="000000"),
        bottom = Side(border_style='thick', color="000000")
    )
    for i in range(start_row, start_row+height):
        for j in range(start_col, start_col+width):
            cell = worksheet[compute_cell(j,i)]
            cell.fill = data_background[i%2]
            cell.font = data_font[0]
            cell.border = data_border
            if (j != start_col):
                if i == start_row+height-1 and is_last_percent:
                    cell.number_format = '0.00 %'
                else:
                    cell.number_format = f'0.00 {LANGUAGE_DICT['currency']}'
                if i == start_row+height-1 and is_last_total:
                    if cell.value < 0.0:
                        cell.font = data_font[2]
                    else:
                        cell.font = data_font[3]
            else:
                cell.font = data_font[1]

def worksheet_table_vertical_background(worksheet, width=2, height=3, start_row=1, start_col=1, is_last_total=False, is_last_col_total=False):
    ## Apply Title Background
    title_background = PatternFill("solid", fgColor="A0A0A0")
    title_font = Font(color="000000", bold=True, size=20)
    title_border = Border(
        left   = Side(border_style='thick', color="000000"),
        right  = Side(border_style='thick', color="000000"),
        top    = Side(border_style='thick', color="000000"),
        bottom = Side(border_style='thick', color="000000")
    )
    for j in range(start_col, start_col+width):
        cell = worksheet[compute_cell(j,start_row)]
        cell.fill = title_background
        cell.font = title_font
        cell.border = title_border
        if (j == start_col):
            worksheet.column_dimensions[get_column(j)].width = 30
        else:
            worksheet.column_dimensions[get_column(j)].width = 24
    ## Apply Values Background
    data_background = [PatternFill("solid", fgColor="F0F0F0"), PatternFill("solid", fgColor="E0E0E0")]
    data_font = [Font(color="000000", bold=False, size=14), Font(color="000000", bold=True, size=14), Font(color="AA0000", bold=False, size=14), Font(color="00AA00", bold=False, size=14)]
    data_border = Border(
        left   = Side(border_style='thick', color="000000"),
        right  = Side(border_style='thick', color="000000"),
        top    = Side(border_style='thin', color="000000"),
        bottom = Side(border_style='thin', color="000000")
    )
    for i in range(start_row+1, start_row+height):
        for j in range(start_col, start_col+width):
            cell = worksheet[compute_cell(j,i)]
            cell.fill = data_background[i%2]
            cell.font = data_font[0]
            cell.border = data_border
            if (j != start_col):
                cell.number_format = f'0.00 {LANGUAGE_DICT['currency']}'
                if j == start_col+width-1 and is_last_col_total:
                    if cell.value < 0.0:
                        cell.font = data_font[2]
                    else:
                        cell.font = data_font[3]
            else:
                cell.font = data_font[1]
    ## Apply total line styling
    if is_last_total:
        total_background = PatternFill("solid", fgColor="C0C0C0")
        total_font = [Font(color="000000", bold=True, size=16), Font(color="AA0000", bold=True, size=16), Font(color="00AA00", bold=True, size=16)]
        total_border = Border(
            left   = Side(border_style='thick', color="000000"),
            right  = Side(border_style='thick', color="000000"),
            top    = Side(border_style='thick', color="000000"),
            bottom = Side(border_style='thick', color="000000")
        )
        for j in range(start_col, start_col+width):
            cell = worksheet[compute_cell(j,start_row+height-1)]
            cell.fill = total_background
            cell.font = total_font[0]
            if j == start_col+width-1 and is_last_col_total:
                if cell.value < 0.0:
                    cell.font = total_font[1]
                elif cell.value == 0.0:
                    cell.font = total_font[0]
                else:
                    cell.font = total_font[2]
            cell.border = total_border
    else:
        last_border = Border(
            left   = Side(border_style='thick', color="000000"),
            right  = Side(border_style='thick', color="000000"),
            top    = Side(border_style='thin', color="000000"),
            bottom = Side(border_style='thick', color="000000")
        )
        for j in range(start_col, start_col+width):
            cell = worksheet[compute_cell(j,start_row+height-1)]
            cell.border = last_border

def worksheet_table_horizontal_background(worksheet, width=2, height=3, start_row=1, start_col=1, is_last_total=0):
    ## Apply Title Background
    title_background = PatternFill("solid", fgColor="A0A0A0")
    title_font = Font(color="000000", bold=True, size=20)
    title_border = Border(
        left   = Side(border_style='thick', color="000000"),
        right  = Side(border_style='thick', color="000000"),
        top    = Side(border_style='thick', color="000000"),
        bottom = Side(border_style='thick', color="000000")
    )
    for j in range(start_col, start_col+width):
        cell = worksheet[compute_cell(j,start_row)]
        cell.fill = title_background
        cell.font = title_font
        cell.border = title_border
        if (j == start_col):
            worksheet.column_dimensions[get_column(j)].width = 30
        else:
            worksheet.column_dimensions[get_column(j)].width = 24
    ## Apply Values Background
    data_background = [PatternFill("solid", fgColor="F0F0F0"), PatternFill("solid", fgColor="E0E0E0")]
    data_font = [Font(color="000000", bold=False, size=14), Font(color="000000", bold=True, size=16), Font(color="AA0000", bold=False, size=14), Font(color="00AA00", bold=False, size=14)]
    data_border = Border(
        left   = Side(border_style='thick', color="000000"),
        right  = Side(border_style='thick', color="000000"),
        top    = Side(border_style='thin', color="000000"),
        bottom = Side(border_style='thin', color="000000")
    )
    for i in range(start_row+1, start_row+height):
        for j in range(start_col, start_col+width):
            cell = worksheet[compute_cell(j,i)]
            cell.fill = data_background[i%2]
            cell.font = data_font[0]
            cell.border = data_border
            if (j != start_col):
                cell.number_format = f'0.00 {LANGUAGE_DICT['currency']}'
            else:
                cell.font = data_font[1]
    ## Apply total line styling
    if is_last_total > 0:
        total_background = PatternFill("solid", fgColor="C0C0C0")
        total_font = [Font(color="000000", bold=True, size=16), Font(color="AA0000", bold=True, size=16), Font(color="00AA00", bold=True, size=16)]
        total_border = Border(
            left   = Side(border_style='thick', color="000000"),
            right  = Side(border_style='thick', color="000000"),
            top    = Side(border_style='thick', color="000000"),
            bottom = Side(border_style='thick', color="000000")
        )
        for i in range(1, is_last_total+1):
            for j in range(start_col, start_col+width):
                cell = worksheet[compute_cell(j,start_row+height-i)]
                cell.fill = total_background
                cell.font = total_font[0]
                if j > start_col:
                    if cell.value < 0.0:
                        cell.font = total_font[1]
                    elif cell.value == 0.0:
                        cell.font = total_font[0]
                    else:
                        cell.font = total_font[2]
                cell.border = total_border
    else:
        last_border = Border(
            left   = Side(border_style='thick', color="000000"),
            right  = Side(border_style='thick', color="000000"),
            top    = Side(border_style='thin', color="000000"),
            bottom = Side(border_style='thick', color="000000")
        )
        for j in range(start_col, start_col+width):
            cell = worksheet[compute_cell(j,start_row+height-1)]
            cell.border = last_border

def set_pie_charts_style(graph:PieChart):
    graph.style = 10
    graph.layout = Layout(
        manualLayout=ManualLayout(
            xMode='edge',
            yMode='edge',
            x = 0.15, y = 0.15,
            w = 0.7 , h = 0.7
        )
    )
    graph.display_blanks = 'gap'
    graph.dataLabels = DataLabelList()
    graph.dataLabels.showPercent = True
    graph.dataLabels.showVal = False
    graph.dataLabels.showSerName = False
    graph.dataLabels.showCatName = False

def set_line_chart_style(graph:LineChart, title="Revenues/Expenses", colors=[]):
    graph.style = 10
    graph.varyColors = False
    graph.layout = Layout(
        manualLayout=ManualLayout(
            xMode='edge',
            yMode='edge',
            x = 0.05, y = 0.15,
            w = 0.9 , h = 0.7
        )
    )
    if (len(colors) != len(graph.series)):
        lines_colors = ["00AA00", "AA0000"]
    else:
        lines_colors = colors
    for i in range(len(graph.series)):
        serie = graph.series[i]
        serie.smooth = False
        serie.graphicalProperties.line.solidFill = lines_colors[i]
        if (len(graph.series) == 1):
            serie.graphicalProperties.line.solidFill = "0000AA"
    graph.title = title
    graph.x_axis.title = LANGUAGE_DICT['month']
    graph.x_axis.delete = False
    graph.y_axis.title = f"{LANGUAGE_DICT['amount']} ({LANGUAGE_DICT['currency']})"
    graph.y_axis.number_format = f'0.00 {LANGUAGE_DICT['currency']}'
    graph.y_axis.delete = False

def generate_anchor(graph, start_col=0, start_row=0, height=1, width=1):
    anchor = TwoCellAnchor()
    anchor._from.col = start_col
    anchor._from.row = start_row
    anchor.to.col = start_col + width
    anchor.to.row = start_row + height
    graph.anchor = anchor

def generate_pie_chart(worksheet, title="Revenus", len_table=1, data_row=2, data_col=3, label_col=2, graph_width=8, graph_height=18, graph_row=1, graph_col=6):
    chart = PieChart()
    label = Reference(worksheet, min_col=label_col, max_col=label_col, min_row=data_row+1, max_row=len_table+data_row-1)
    data  = Reference(worksheet, min_col=data_col, max_col=data_col, min_row=data_row, max_row=len_table+data_row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(label)
    set_pie_charts_style(chart)
    chart.title = title
    generate_anchor(chart, start_col=graph_col, start_row=graph_row, height=graph_height, width=graph_width)
    worksheet.add_chart(chart)

def generate_line_chart(worksheet, title="Revenus", len_table=1, nb_lines=1, data_row=2, data_col=3, label_row=2, graph_width=8, graph_height=18, graph_row=1, graph_col=6, colors=[]):
    chart = LineChart()
    label = Reference(worksheet, min_col=data_col+1, max_col=data_col+len_table+1, min_row=label_row, max_row=label_row)
    data  = Reference(worksheet, min_col=data_col, max_col=data_col+len_table, min_row=data_row, max_row=data_row+nb_lines-1)
    chart.add_data(data, from_rows=True, titles_from_data=True)
    chart.set_categories(label)
    set_line_chart_style(chart, title=title, colors=colors)
    generate_anchor(chart, start_col=graph_col, start_row=graph_row, height=graph_height, width=graph_width)
    worksheet.add_chart(chart)